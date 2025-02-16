import os
import requests
from bs4 import BeautifulSoup
import random
from openpyxl import load_workbook, Workbook

def clear_terminal():
    os.system('cls' if os.name == 'nt' else 'clear')
    #in windows os.name == 'nt' and clear command in windows is 'cls' and in other OS it is 'Clear'

def fetch_definition(word):
    #Fetch word definition from Merriam-Webster.
    url = f'https://www.merriam-webster.com/dictionary/{word}'
    response = requests.get(url)
    
    if response.status_code != 200:
        print("Error fetching the definition.")
        return None, None
    
    soup = BeautifulSoup(response.content, 'html.parser')
    word_element = soup.select_one('h1.hword')
    definitions = soup.select('span.dtText')
    
    if word_element and definitions:
        return word_element.get_text(strip=True), definitions[0].get_text(strip=True)[1:]
    return None, None

def load_or_create_workbook(filename='bank.xlsx'):
    #Load workbook if exists, otherwise create a new one without overwriting.
    if os.path.exists(filename):
        return load_workbook(filename)
    
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["Word", "Definition"])  # Adds a header
    workbook.save(filename)
    return workbook


def get_word_count():
    #Return total number of words in the Excel sheet. useful for randomly checking a number
    workbook = load_or_create_workbook()
    worksheet = workbook.active
    return worksheet.max_row

def get_random_word():
    #Select a random word and its definition from the Excel sheet.
    workbook = load_or_create_workbook()
    worksheet = workbook.active
    row = random.randint(1, get_word_count())
    return worksheet.cell(row=row, column=1).value, worksheet.cell(row=row, column=2).value

def add_word_to_excel(word, definition):
    #Add a new word and its definition to the Excel sheet if it doesn't already exist.
    workbook = load_or_create_workbook()
    worksheet = workbook.active
    
    existing_words = {row[0] for row in worksheet.iter_rows(min_col=1, max_col=1, values_only=True) if row[0]}
    if word in existing_words:
        print(f"'{word}' already exists in the database.")
        return False
    
    worksheet.append([word, definition])
    workbook.save('bank.xlsx')
    return True

def update_word_bank():
    #Update the word bank by fetching new words and its definitions.
    while True:
        word = input("Enter the word to update (or 'e' to exit): ").strip()
        if word.lower() == 'e':
            print("Exiting update mode.")
            break
        
        fetched_word, definition = fetch_definition(word)
        if fetched_word and definition and add_word_to_excel(fetched_word, definition):
            print("Word added successfully!")

def play_game():
    #Start the word guessing game.
    word, definition = get_random_word()
    if not word or not definition:
        print("No words available in the database. Please update first.")
        return
    
    print(f"Hint: {definition}")
    for _ in range(3):
        #Gives the player 3 chances to guess the word.
        guess = input("Guess the word: ").strip().lower()
        if guess == word.lower():
            print("Congratulations! You won the game!")
            return
    print("Better luck next time!!!")

def main():
    #Main program loop.
    workbook = load_or_create_workbook()
    
    while True:
        # While has True in place of condition, ensuring that Unless the Exit command is given the game will start once again.
        choice = input("Choose an option - 'Play', 'Update', or 'Exit': ").strip().lower()
        if choice in ('exit', 'e'):
            clear_terminal()
            print("!!! Game Closed !!!")
            break
        elif choice in ('update', 'u'):
            update_word_bank()
        elif choice in ('play', 'p'):
            play_game()
        else:
            print("Invalid input. Please try again.")

if __name__ == '__main__':
    #Enures that the main function runs only when the program is executed directly.
    main()
